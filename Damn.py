#!/usr/bin/env python3
"""
Ads.txt / App-Ads.txt Crawler  v5
  - Output domain order ALWAYS matches the order you supplied (no more
    random ordering from threaded crawling).
  - Network Finder now returns a compact 3-column table:
        Domain | Found (Yes/No/Error) | Details
    instead of one column per seller ID.
  - match_fields selector (2 = domain+ID default, 3 = +relation, 4 = +tag)
  - Smarter fetch: retry, HTML-page detection, detailed error labels
"""

import io
import os
import re
import tempfile
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from typing import Dict, List, Optional
import pandas as pd
import requests
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

# ── Persistent saved-config storage ─────────────────────────────────────────
# Plain text files saved next to this script. Each holds raw text exactly as
# pasted/uploaded, so the existing parse_* functions can be reused unchanged.
#
# This block is intentionally defensive: it must NEVER raise at import time,
# because that would crash the whole app on every rerun. If the script's own
# directory isn't writable, fall back to the system temp dir. If even that
# fails, persistence is disabled (PERSISTENCE_OK = False) but the rest of the
# app still works exactly as before - saved_exists()/load_saved() just behave
# as "nothing saved".

PERSISTENCE_OK = False
PERSISTENCE_WARNING = ""
SAVED_DIR = ""


def _try_dir(path: str) -> bool:
    """Return True if `path` exists (or can be created) and is writable."""
    try:
        os.makedirs(path, exist_ok=True)
        test_path = os.path.join(path, ".write_test")
        with open(test_path, "w", encoding="utf-8") as f:
            f.write("ok")
        os.remove(test_path)
        return True
    except OSError:
        return False


try:
    _script_dir = os.path.dirname(os.path.abspath(__file__))
except NameError:
    _script_dir = os.getcwd()

_candidate_primary = os.path.join(_script_dir, "saved_config")
_candidate_fallback = os.path.join(tempfile.gettempdir(), "ads_txt_crawler_saved_config")

if _try_dir(_candidate_primary):
    SAVED_DIR = _candidate_primary
    PERSISTENCE_OK = True
elif _try_dir(_candidate_fallback):
    SAVED_DIR = _candidate_fallback
    PERSISTENCE_OK = True
    PERSISTENCE_WARNING = (
        f"Couldn't write to `{_candidate_primary}` (read-only?), so saved "
        f"config is using a temporary folder instead. It will work for now "
        f"but may not survive an app restart/redeploy."
    )
else:
    PERSISTENCE_WARNING = (
        "Persistent storage isn't available in this environment (no writable "
        "folder found). The 💾 Saved Config tab is disabled for now, but the "
        "rest of the app works as usual."
    )

SAVED_FILES = {
    "domains":      os.path.join(SAVED_DIR, "domains.txt") if SAVED_DIR else "",
    "lines_ads":    os.path.join(SAVED_DIR, "lines_ads.txt") if SAVED_DIR else "",
    "lines_appads": os.path.join(SAVED_DIR, "lines_appads.txt") if SAVED_DIR else "",
    "seller_ids":   os.path.join(SAVED_DIR, "seller_ids.txt") if SAVED_DIR else "",
}

SAVED_LABELS = {
    "domains":      "Domains",
    "lines_ads":    "Lines to check — ads.txt",
    "lines_appads": "Lines to check — app-ads.txt",
    "seller_ids":   "Seller IDs (Network Finder / Combined Export)",
}


def load_saved(key: str) -> str:
    if not PERSISTENCE_OK:
        return ""
    path = SAVED_FILES.get(key, "")
    if path and os.path.exists(path):
        try:
            with open(path, "r", encoding="utf-8") as f:
                return f.read()
        except OSError:
            return ""
    return ""


def save_saved(key: str, content: str) -> bool:
    """Returns True on success. Never raises."""
    if not PERSISTENCE_OK:
        return False
    path = SAVED_FILES.get(key, "")
    if not path:
        return False
    try:
        with open(path, "w", encoding="utf-8") as f:
            f.write(content)
        return True
    except OSError:
        return False


def clear_saved(key: str) -> bool:
    if not PERSISTENCE_OK:
        return False
    path = SAVED_FILES.get(key, "")
    try:
        if path and os.path.exists(path):
            os.remove(path)
        return True
    except OSError:
        return False


def saved_exists(key: str) -> bool:
    if not PERSISTENCE_OK:
        return False
    return bool(load_saved(key).strip())


def safe_rerun() -> None:
    """st.rerun() compatibility shim for older Streamlit versions."""
    if hasattr(st, "rerun"):
        st.rerun()
    else:
        st.experimental_rerun()


# ── Page config ───────────────────────────────────────────────────────────────

st.set_page_config(
    page_title="Ads.txt Crawler",
    page_icon="🔍",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── CSS ───────────────────────────────────────────────────────────────────────

st.markdown("""
<style>
[data-testid="stSidebar"] {
    background:linear-gradient(180deg,#0f0c29,#302b63,#24243e);
}
[data-testid="stSidebar"] label,
[data-testid="stSidebar"] p,
[data-testid="stSidebar"] span,
[data-testid="stSidebar"] div { color:#dde1ff !important; }
[data-testid="stSidebar"] .stButton>button {
    background:linear-gradient(135deg,#667eea,#764ba2)!important;
    color:white!important;border:none!important;
    border-radius:10px!important;font-weight:700!important;
}
[data-testid="stSidebar"] hr{border-color:#444!important;}

.hero{
    background:linear-gradient(135deg,#667eea 0%,#764ba2 100%);
    padding:1.3rem 1.8rem;border-radius:12px;color:white;
    margin-bottom:1.2rem;box-shadow:0 4px 20px rgba(102,126,234,.3);
}
.hero h1{margin:0;font-size:1.8rem;}
.hero p {margin:.3rem 0 0;opacity:.9;font-size:.95rem;}

.ft-pill{
    display:inline-block;padding:2px 12px;border-radius:99px;
    font-size:.9rem;font-weight:700;vertical-align:middle;margin-left:8px;
}
.ft-appads{background:#667eea;color:#fff;}
.ft-ads   {background:#f59e0b;color:#fff;}

.metric-row{display:flex;gap:.8rem;margin:1rem 0;flex-wrap:wrap;}
.tile{
    flex:1;min-width:110px;background:white;
    border-radius:10px;padding:.9rem 1rem;
    box-shadow:0 2px 8px rgba(0,0,0,.07);
    border-top:4px solid #667eea;
}
.tile.green{border-top-color:#22c55e;}
.tile.red  {border-top-color:#ef4444;}
.tile.amber{border-top-color:#f59e0b;}
.tile-lbl{font-size:.72rem;color:#6b7280;font-weight:600;letter-spacing:.04em;}
.tile-val{font-size:1.8rem;font-weight:800;color:#111;}

.mode-tag{
    display:inline-block;background:#f0f4ff;color:#3730a3;
    border:1px solid #c7d2fe;border-radius:6px;
    padding:2px 10px;font-size:.8rem;font-weight:600;margin-bottom:.5rem;
}
.order-note{
    font-size:.8rem;color:#6b7280;font-style:italic;margin-top:.3rem;
}
</style>
""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
# CORE  LOGIC
# ══════════════════════════════════════════════════════════════════════════════

def normalize_domain(raw: str) -> str:
    raw = raw.strip()
    raw = re.sub(r"^https?://", "", raw, flags=re.IGNORECASE)
    return raw.split("/")[0].split("?")[0].split("#")[0].lower().strip()


def parse_ads_line(raw: str) -> Optional[Dict]:
    raw = raw.strip()
    if not raw or raw.startswith("#"):
        return None
    raw = raw.split("#")[0].strip()
    parts = [p.strip() for p in raw.split(",")]
    if len(parts) < 3:
        return None
    return {
        "domain":    parts[0].lower(),
        "seller_id": parts[1],
        "relation":  parts[2].upper(),
        "tag":       parts[3].lower() if len(parts) > 3 else "",
    }


def lines_match(entry: Dict, query: Dict, match_fields: int = 2) -> bool:
    """
    match_fields controls how many fields are compared:
      2 -> domain + seller_id   (default - most lenient)
      3 -> + relation
      4 -> + tag (only when query includes a tag)
    Domain is ALWAYS compared (it defines the ad network).
    """
    if entry["domain"] != query["domain"]:
        return False
    if match_fields >= 2 and entry["seller_id"].lower() != query["seller_id"].lower():
        return False
    if match_fields >= 3 and entry["relation"] != query["relation"]:
        return False
    if match_fields >= 4 and query["tag"] and entry["tag"].lower() != query["tag"]:
        return False
    return True


def fetch_file(domain: str, file_type: str, timeout: int) -> Dict:
    """
    Fetch ads.txt / app-ads.txt with:
      - HTTPS -> HTTP fallback
      - 1 automatic retry on transient connection / server errors
      - HTML-page detection (catches servers that return 200 + error HTML)
      - Specific error labels (timeout / DNS / SSL / HTTP code / empty / HTML)
    """
    headers = {
        "User-Agent": "AdsTxtCrawler/5.0",
        "Accept": "text/plain, text/*, */*",
    }
    last_error = "Connection failed"
    max_attempts = 2  # 1 initial + 1 retry

    for attempt in range(max_attempts):
        if attempt > 0:
            time.sleep(0.8)  # brief back-off before retry

        for scheme in ("https", "http"):
            url = f"{scheme}://{domain}/{file_type}"
            try:
                r = requests.get(
                    url, headers=headers, timeout=timeout, allow_redirects=True
                )

                if r.status_code == 200:
                    ctype   = r.headers.get("Content-Type", "")
                    content = r.text.strip()

                    if not content:
                        last_error = "Empty file (200 OK but no content)"
                        continue

                    if "html" in ctype.lower() or re.match(
                        r"<\s*(!doctype|html)", content[:40], re.I
                    ):
                        last_error = "HTML page returned (file probably missing)"
                        continue

                    return {"ok": True, "url": url, "text": r.text, "error": ""}

                elif r.status_code == 404:
                    return {
                        "ok": False, "url": url, "text": "",
                        "error": "File not found (404)",
                    }
                elif r.status_code in (401, 403):
                    return {
                        "ok": False, "url": url, "text": "",
                        "error": f"Access denied (HTTP {r.status_code})",
                    }
                elif r.status_code >= 500:
                    last_error = f"Server error (HTTP {r.status_code})"
                else:
                    last_error = f"Unexpected HTTP {r.status_code}"

            except requests.exceptions.Timeout:
                last_error = f"Request timed out ({timeout}s)"
                continue  # try next scheme (https -> http) before giving up
            except requests.exceptions.SSLError:
                last_error = "SSL / certificate error"
            except requests.exceptions.ConnectionError as exc:
                msg = str(exc).lower()
                if any(k in msg for k in ("getaddrinfo", "name or service", "nodename", "nxdomain")):
                    return {
                        "ok": False, "url": url, "text": "",
                        "error": "DNS lookup failed (domain not resolved)",
                    }
                elif "connection refused" in msg:
                    return {
                        "ok": False, "url": url, "text": "",
                        "error": "Connection refused by server",
                    }
                else:
                    last_error = "Network connection error"
            except Exception as exc:
                last_error = f"Unexpected error: {str(exc)[:60]}"

    return {
        "ok": False,
        "url": f"https://{domain}/{file_type}",
        "text": "",
        "error": last_error,
    }


def parse_domain_list(text: str) -> List[str]:
    """Returns domains in the EXACT order they appear, de-duplicated."""
    seen, result = set(), []
    for line in text.splitlines():
        line = line.strip()
        if not line or line.startswith("#"):
            continue
        d = normalize_domain(line)
        if d and d not in seen:
            seen.add(d)
            result.append(d)
    return result


def parse_query_list(text: str) -> List[Dict]:
    result, seen = [], set()
    for line in text.splitlines():
        q = parse_ads_line(line)
        if q:
            key = _qkey(q)
            if key not in seen:
                seen.add(key)
                result.append(q)
    return result


def parse_seller_ids(text: str) -> List[str]:
    """One seller ID per line (or comma-separated). Strips duplicates, keeps order."""
    ids, seen = [], set()
    for line in text.splitlines():
        for part in line.split(","):
            sid = part.strip()
            if sid and not sid.startswith("#") and sid not in seen:
                seen.add(sid)
                ids.append(sid)
    return ids


def _qkey(q: Dict) -> str:
    parts = [q["domain"], q["seller_id"], q["relation"]]
    if q["tag"]:
        parts.append(q["tag"])
    return ", ".join(parts)


# ── Crawl workers ─────────────────────────────────────────────────────────────

def crawl_lines(
    domain: str,
    queries: List[Dict],
    timeout: int,
    file_type: str,
    match_fields: int,
) -> List[Dict]:
    """Tab 1: one result row per (domain, query)."""
    fetch = fetch_file(domain, file_type, timeout)
    rows  = []
    if fetch["ok"]:
        entries = [
            e for ln in fetch["text"].splitlines()
            if (e := parse_ads_line(ln)) is not None
        ]
        for q in queries:
            match = next(
                (e for e in entries if lines_match(e, q, match_fields)), None
            )
            rows.append({
                "Domain": domain,
                "Line":   _qkey(q),
                "Found":  "Yes" if match else "No",
                "_ok":    True,
                "_err":   "",
            })
    else:
        for q in queries:
            rows.append({
                "Domain": domain,
                "Line":   _qkey(q),
                "Found":  "Error",
                "_ok":    False,
                "_err":   fetch["error"],
            })
    return rows


def crawl_network_all(domain: str, network: str, relation: str,
                       timeout: int, file_type: str) -> Dict:
    """
    Tab 2 - Find-All mode.
    Returns ONE row per domain: Domain | Found | Details
      Found   = Yes  -> at least one matching entry exists
      Details = list of every matching seller ID found
    """
    fetch = fetch_file(domain, file_type, timeout)
    if not fetch["ok"]:
        return {"Domain": domain, "Found": "Error",
                "Details": fetch["error"], "_ok": False, "_err": fetch["error"]}

    entries = [
        e for ln in fetch["text"].splitlines()
        if (e := parse_ads_line(ln)) is not None
    ]
    matching = [
        e for e in entries
        if e["domain"] == network.lower()
        and (not relation or e["relation"] == relation.upper())
    ]
    ids = [e["seller_id"] for e in matching]

    if ids:
        unit = "entry" if len(ids) == 1 else "entries"
        details = f"{len(ids)} {unit} found - " + ", ".join(ids)
        found = "Yes"
    else:
        details = "No matching entries"
        found = "No"

    return {"Domain": domain, "Found": found, "Details": details,
            "_ok": True, "_err": ""}


def crawl_network_specific(domain: str, network: str, relation: str,
                            seller_ids: List[str], timeout: int,
                            file_type: str) -> Dict:
    """
    Tab 2 - Specific Seller IDs mode.
    For ONE domain, check whether ANY of the supplied seller_ids appear
    paired with `network` (and `relation`, if given).
    Returns ONE row: Domain | Found | Details
      Found   = Yes  -> at least one of the supplied seller IDs matched
      Details = which seller ID(s) matched, or "0 of N matched"
    """
    fetch = fetch_file(domain, file_type, timeout)
    if not fetch["ok"]:
        return {"Domain": domain, "Found": "Error",
                "Details": fetch["error"], "_ok": False, "_err": fetch["error"]}

    entries = [
        e for ln in fetch["text"].splitlines()
        if (e := parse_ads_line(ln)) is not None
    ]

    # All seller IDs (lower-cased) that exist for this network (+ relation if given)
    net_entries = [
        e for e in entries
        if e["domain"] == network.lower()
        and (not relation or e["relation"] == relation.upper())
    ]
    available = {e["seller_id"].lower() for e in net_entries}

    matched = [sid for sid in seller_ids if sid.lower() in available]
    n = len(seller_ids)

    if matched:
        details = f"Matched {len(matched)} of {n} - " + ", ".join(matched)
        found = "Yes"
    else:
        details = f"0 of {n} matched"
        found = "No"

    return {"Domain": domain, "Found": found, "Details": details,
            "_ok": True, "_err": ""}


def crawl_combined(
    domain: str,
    network: str,
    relation: str,
    seller_ids: List[str],
    queries: List[Dict],
    match_fields: int,
    timeout: int,
    file_type: str,
) -> Dict:
    """
    Tab 3 – Combined Export.
    Fetches the file ONCE per domain, then runs:
      • Network check  (find-all or specific seller IDs)
      • All line checks
    Returns a single flat dict per domain.
    """
    fetch = fetch_file(domain, file_type, timeout)
    row: Dict = {"Domain": domain}

    if not fetch["ok"]:
        row["Network Found"]   = "Error"
        row["Network Details"] = fetch["error"]
        for q in queries:
            row[_qkey(q)] = "Error"
        return row

    entries = [
        e for ln in fetch["text"].splitlines()
        if (e := parse_ads_line(ln)) is not None
    ]

    # ── Network check ─────────────────────────────────────────────────────────
    net_entries = [
        e for e in entries
        if e["domain"] == network.lower()
        and (not relation or e["relation"] == relation.upper())
    ]

    if seller_ids:                        # Specific-IDs mode
        available = {e["seller_id"].lower() for e in net_entries}
        matched   = [sid for sid in seller_ids if sid.lower() in available]
        n         = len(seller_ids)
        if matched:
            row["Network Found"]   = "Yes"
            row["Network Details"] = f"Matched {len(matched)} of {n}: " + ", ".join(matched)
        else:
            row["Network Found"]   = "No"
            row["Network Details"] = f"0 of {n} matched"
    else:                                  # Find-All mode
        ids = [e["seller_id"] for e in net_entries]
        if ids:
            unit = "entry" if len(ids) == 1 else "entries"
            row["Network Found"]   = "Yes"
            row["Network Details"] = f"{len(ids)} {unit}: " + ", ".join(ids)
        else:
            row["Network Found"]   = "No"
            row["Network Details"] = "No matching entries"

    # ── Line checks ───────────────────────────────────────────────────────────
    for q in queries:
        match = next((e for e in entries if lines_match(e, q, match_fields)), None)
        row[_qkey(q)] = "Yes" if match else "No"

    return row


def run_parallel(domain_list: List[str], task_fn, workers: int) -> List:
    """
    Runs task_fn(domain) for every domain in parallel.
    NOTE: results come back in COMPLETION order (not submission order) -
    callers MUST reorder by domain_list afterwards (see reorder_*).
    """
    all_results, done, total = [], 0, len(domain_list)
    prog = st.progress(0.0, text=f"0 / {total} domains")
    t0 = time.time()
    with ThreadPoolExecutor(max_workers=workers) as pool:
        fmap = {pool.submit(task_fn, d): d for d in domain_list}
        for fut in as_completed(fmap):
            dom = fmap[fut]
            try:
                res = fut.result()
                (all_results.extend if isinstance(res, list) else all_results.append)(res)
            except Exception as exc:
                # Safety net - should rarely trigger since fetch_file()
                # already catches everything internally.
                all_results.append({
                    "Domain": dom, "Found": "Error",
                    "Details": f"Internal error: {str(exc)[:80]}",
                    "Line": "", "_ok": False,
                    "_err": f"Internal error: {str(exc)[:80]}",
                })
            done += 1
            prog.progress(
                done / total,
                text=f"🔄 {done} / {total} - {dom} ({round(time.time()-t0,1)}s)",
            )
    prog.progress(1.0, text=f"✅ Finished in {round(time.time()-t0,1)}s")
    return all_results


# ── Order-preservation helpers ────────────────────────────────────────────────

def reorder_rows_by_domain(rows: List[Dict], domain_order: List[str]) -> List[Dict]:
    """
    Reorder a flat list of per-domain dicts (one row per domain) so the
    result follows domain_order exactly. Any domain missing a result
    (should not normally happen) gets a placeholder error row.
    """
    by_domain: Dict[str, Dict] = {}
    for r in rows:
        by_domain.setdefault(r["Domain"], r)

    ordered = []
    for d in domain_order:
        if d in by_domain:
            ordered.append(by_domain[d])
        else:
            ordered.append({
                "Domain": d, "Found": "Error",
                "Details": "No result returned",
                "_ok": False, "_err": "No result returned",
            })
    return ordered


def build_pivot(raw_rows: List[Dict], line_order: List[str],
               domain_order: List[str]) -> pd.DataFrame:
    """
    Build the Domain x Line pivot table.
    Rows follow domain_order, columns follow line_order - both exactly
    as supplied by the user, regardless of the order threads finished in.
    """
    df = pd.DataFrame(raw_rows)
    pivot = df.pivot_table(
        index="Domain", columns="Line", values="Found", aggfunc="first"
    )
    cols = [k for k in line_order if k in pivot.columns]
    pivot = pivot[cols]
    pivot = pivot.reindex(domain_order)   # <- enforce user-supplied order
    pivot = pivot.fillna("Error")          # any domain with no rows -> Error
    pivot.index.name = pivot.columns.name = None
    out = pivot.copy()
    out.insert(0, "Crawled Domain", pivot.index)
    return out.reset_index(drop=True)


def reorder_flat(df_long: pd.DataFrame, domain_order: List[str],
                line_order: List[str]) -> pd.DataFrame:
    """Order a flat (Domain, Line, Found) table by domain_order then line_order."""
    d_rank = {d: i for i, d in enumerate(domain_order)}
    l_rank = {l: i for i, l in enumerate(line_order)}
    out = df_long.copy()
    out["_dr"] = out["Domain"].map(d_rank).fillna(len(domain_order))
    out["_lr"] = out["Line"].map(l_rank).fillna(len(line_order))
    out = out.sort_values(["_dr", "_lr"], kind="stable").drop(columns=["_dr", "_lr"])
    return out.reset_index(drop=True)


def build_combined_excel(
    rows: List[Dict],
    network_label: str,
    line_keys: List[str],
) -> bytes:
    """
    Build a colour-coded Excel workbook for the Combined Export tab.
    Sheet layout:
        Col A  : Domain
        Col B  : Network Found  (Yes / No / Error)
        Col C  : Network Details (seller IDs / match info / error text)
        Col D+ : one column per ads-txt line checked  (Yes / No / Error)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Combined Results"

    # ── Header row ────────────────────────────────────────────────────────────
    net_col  = f"Network Found\n({network_label})"
    headers  = ["Domain", net_col, "Network Details"] + line_keys
    ws.append(headers)

    hdr_fill = PatternFill(start_color="4B5EAA", end_color="4B5EAA", fill_type="solid")
    hdr_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    hdr_aln  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for cell in ws[1]:
        cell.fill      = hdr_fill
        cell.font      = hdr_font
        cell.alignment = hdr_aln
    ws.row_dimensions[1].height = 40

    # ── Value colour fills & fonts ────────────────────────────────────────────
    YES_FILL  = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    NO_FILL   = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    ERR_FILL  = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    YES_FONT  = Font(color="166534", bold=True, name="Arial", size=10)
    NO_FONT   = Font(color="991B1B", bold=True, name="Arial", size=10)
    ERR_FONT  = Font(color="92400E", bold=True, name="Arial", size=10)
    BASE_FONT = Font(name="Arial", size=10)
    BASE_ALN  = Alignment(vertical="center", wrap_text=True)
    CTR_ALN   = Alignment(horizontal="center", vertical="center")

    fill_map = {"Yes": YES_FILL, "No": NO_FILL, "Error": ERR_FILL}
    font_map = {"Yes": YES_FONT, "No": NO_FONT,  "Error": ERR_FONT}

    # ── Data rows ─────────────────────────────────────────────────────────────
    for row in rows:
        vals = [
            row.get("Domain",          ""),
            row.get("Network Found",   ""),
            row.get("Network Details", ""),
        ] + [row.get(k, "Error") for k in line_keys]
        ws.append(vals)

        r = ws.max_row
        ws.row_dimensions[r].height = 18
        for col_idx, val in enumerate(vals, 1):
            cell           = ws.cell(r, col_idx)
            cell.font      = font_map.get(val, BASE_FONT)
            cell.alignment = CTR_ALN if col_idx != 3 else BASE_ALN
            if val in fill_map:
                cell.fill = fill_map[val]

    # ── Column widths ─────────────────────────────────────────────────────────
    for i, col in enumerate(ws.columns):
        raw_max = max((len(str(cell.value or "")) for cell in col), default=10)
        if i == 0:        # Domain
            width = min(raw_max + 4, 40)
        elif i == 2:      # Network Details (can be long)
            width = min(raw_max + 4, 60)
        else:             # Found cols + line cols
            width = min(raw_max + 4, 45)
        ws.column_dimensions[col[0].column_letter].width = width

    # ── Freeze header row + domain column ─────────────────────────────────────
    ws.freeze_panes = "B2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ── Display helpers ───────────────────────────────────────────────────────────

def _style(val):
    if val == "Yes":   return "background-color:#dcfce7;color:#166534;font-weight:700"
    if val == "No":    return "background-color:#fee2e2;color:#991b1b;font-weight:700"
    if val == "Error": return "background-color:#fef3c7;color:#92400e;font-weight:700"
    return ""


def styled_df(df: pd.DataFrame, cols):
    try:
        return df.style.map(_style, subset=cols)
    except AttributeError:
        return df.style.applymap(_style, subset=cols)


def tiles_html(items):  # items: [(label, value, css_class), ...]
    parts = [
        f'<div class="tile {c}"><div class="tile-lbl">{l}</div>'
        f'<div class="tile-val">{v}</div></div>'
        for l, v, c in items
    ]
    return f'<div class="metric-row">{"".join(parts)}</div>'


def stale_warn(last_ft, current_ft):
    if last_ft and last_ft != current_ft:
        st.warning(
            f"Showing results for **{last_ft}**. "
            f"File type is now **{current_ft}** - re-run to refresh.",
            icon="⚠️",
        )


# ── Session state ─────────────────────────────────────────────────────────────

for _k in ("std_raw", "std_keys", "std_ft", "std_domains",
           "finder_df", "finder_meta", "finder_mode",
           "finder_domains", "finder_nsids",
           "comb_rows", "comb_meta", "comb_line_keys",
           "comb_domains", "comb_nsids"):
    if _k not in st.session_state:
        st.session_state[_k] = None


# ══════════════════════════════════════════════════════════════════════════════
# SIDEBAR
# ══════════════════════════════════════════════════════════════════════════════

with st.sidebar:
    st.markdown("## ⚙️ Settings")

    # 1 -- File type ────────────────────────────────────────────────────────────
    st.markdown("**File type**")
    file_type = st.radio(
        "ft", ["app-ads.txt", "ads.txt"],
        index=0, horizontal=True,
        label_visibility="collapsed",
    )
    ft_cls = "ft-appads" if file_type == "app-ads.txt" else "ft-ads"
    st.markdown(
        f'Crawling: <span class="ft-pill {ft_cls}">{file_type}</span>',
        unsafe_allow_html=True,
    )
    st.divider()

    # 2 -- Domains ────────────────────────────────────────────────────────────
    st.markdown("**Domains**")
    st.caption("Order is preserved in all results.")
    if saved_exists("domains"):
        st.caption("📌 Loaded from Saved Config — edit below or change in 💾 Saved Config tab.")
    dm = st.radio("dm", ["📁 Upload", "✏️ Paste"], horizontal=True,
                  label_visibility="collapsed", key="dm")
    domains_raw = ""
    if dm == "📁 Upload":
        up = st.file_uploader("domains.txt", type="txt", key="dup",
                              label_visibility="collapsed")
        if up:
            domains_raw = up.read().decode("utf-8", errors="replace")
        else:
            domains_raw = load_saved("domains")
    else:
        domains_raw = st.text_area(
            "Domains", value=load_saved("domains"), key="dtxt", height=130,
            label_visibility="collapsed",
            placeholder="apps.mxplayer.in\nbattleprime.com\n…",
        )
    domain_list = parse_domain_list(domains_raw) if domains_raw.strip() else []
    if domain_list:
        st.caption(f"✔ {len(domain_list)} domain(s), in this order:")
        with st.expander("Preview order", expanded=False):
            st.write(", ".join(f"{i+1}. {d}" for i, d in enumerate(domain_list)))
    st.divider()

    # 3 -- Match fields ───────────────────────────────────────────────────────
    st.markdown("**Fields to match** *(Line Checker)*")
    match_fields = st.radio(
        "mf",
        options=[2, 3, 4],
        index=0,
        horizontal=True,
        label_visibility="collapsed",
        format_func=lambda x: {
            2: "2 - Domain + ID",
            3: "3 - + Relation",
            4: "4 - + Tag",
        }[x],
    )
    st.caption({
        2: "Match on **Ad Network domain** and **Seller ID** only.",
        3: "Also require **Relation** (RESELLER / DIRECT) to match.",
        4: "Also require **Tag** (certification ID) to match.",
    }[match_fields])
    st.divider()

    # 4 -- Crawl settings ─────────────────────────────────────────────────────
    st.markdown("**Crawl settings**")
    workers = st.slider("Parallel workers", 1, 30, 8)
    timeout = st.slider("Timeout per domain (s)", 5, 60, 12,
                        help="Higher timeout -> fewer false 'Error' results on slow domains.")


# ══════════════════════════════════════════════════════════════════════════════
# HEADER
# ══════════════════════════════════════════════════════════════════════════════

pill = f'<span class="ft-pill {ft_cls}">{file_type}</span>'
st.markdown(
    f'<div class="hero">'
    f'<h1>🔍 Ads.txt Crawler {pill}</h1>'
    f'<p>'
    f'<b>Line Checker</b> - verify specific entries; pivot table (domains x lines).<br>'
    f'<b>Network Finder</b> - find all, or check specific seller IDs, for one ad network. '
    f'Compact 3-column output: Domain | Found | Details.'
    f'</p></div>',
    unsafe_allow_html=True,
)


# ══════════════════════════════════════════════════════════════════════════════
# TABS
# ══════════════════════════════════════════════════════════════════════════════

tab1, tab2, tab3, tab4 = st.tabs([
    "📋  Line Checker", "🔎  Network Finder", "📊  Combined Export", "💾  Saved Config",
])


# ─────────────────────────────────────────────────────────────────────────────
# TAB 1 - Line Checker
# ─────────────────────────────────────────────────────────────────────────────

with tab1:

    st.markdown(
        f"Checking against **`{file_type}`** · "
        f"matching **{match_fields} field(s)**: "
        + {2: "domain + seller ID", 3: "domain + seller ID + relation",
           4: "domain + seller ID + relation + tag"}[match_fields]
    )

    lines_save_key = "lines_ads" if file_type == "ads.txt" else "lines_appads"
    if saved_exists(lines_save_key):
        st.caption(
            f"📌 Loaded saved lines for **{file_type}** "
            f"— edit below or change in 💾 Saved Config tab."
        )

    lm = st.radio("lm", ["📁 Upload lines.txt", "✏️ Paste lines"],
                  horizontal=True, label_visibility="collapsed", key="lm")
    lines_raw = ""
    if lm == "📁 Upload lines.txt":
        lup = st.file_uploader(
            "lines.txt", type="txt", key="lup",
            label_visibility="collapsed",
            help="One entry per line: domain, seller_id, relation[, tag]",
        )
        if lup:
            lines_raw = lup.read().decode("utf-8", errors="replace")
        else:
            lines_raw = load_saved(lines_save_key)
    else:
        lines_raw = st.text_area(
            "Lines to check", value=load_saved(lines_save_key), key="ltxt", height=150,
            label_visibility="collapsed",
            placeholder=(
                "google.com, pub-6968738577620513, RESELLER, f08c47fec0942fa0\n"
                "smartadserver.com, 5427, RESELLER, 060d053dcf45cbf3\n"
                "video.unrulymedia.com, 906189653, RESELLER"
            ),
        )

    query_list = parse_query_list(lines_raw) if lines_raw.strip() else []
    if query_list:
        st.caption(f"✔ {len(query_list)} line(s) loaded")

    if st.button(
        f"🚀 Run Line Check  [{file_type}]",
        type="primary", key="run_std", use_container_width=True,
    ):
        if not domain_list:
            st.error("Add domains in the sidebar first.")
        elif not query_list:
            st.error("Provide at least one line to check.")
        else:
            _ql, _to, _ft, _mf = query_list[:], timeout, file_type, match_fields
            _domains = domain_list[:]
            raw = run_parallel(
                _domains,
                lambda d: crawl_lines(d, _ql, _to, _ft, _mf),
                workers,
            )
            st.session_state.std_raw     = raw
            st.session_state.std_keys    = [_qkey(q) for q in _ql]
            st.session_state.std_ft      = _ft
            st.session_state.std_domains = _domains

    # ── Results ───────────────────────────────────────────────────────────────
    if st.session_state.std_raw:
        stale_warn(st.session_state.std_ft, file_type)

        raw_rows     = st.session_state.std_raw
        ordered_keys = st.session_state.std_keys or []
        domain_order = st.session_state.std_domains or domain_list
        df_long      = pd.DataFrame(raw_rows)

        yes_n   = (df_long["Found"] == "Yes").sum()
        no_n    = (df_long["Found"] == "No").sum()
        err_n   = (df_long["Found"] == "Error").sum()
        total_n = len(df_long)

        st.markdown(tiles_html([
            ("TOTAL",      f"{total_n:,}",                                       ""),
            ("YES",        f"{yes_n:,}",                                         "green"),
            ("NO",         f"{no_n:,}",                                          "red"),
            ("ERRORS",     f"{err_n:,}",                                         "amber"),
            ("FOUND RATE", f"{round(yes_n/total_n*100,1) if total_n else 0}%",   ""),
        ]), unsafe_allow_html=True)

        try:
            disp = build_pivot(raw_rows, ordered_keys, domain_order)
            val_cols = [c for c in disp.columns if c != "Crawled Domain"]
            st.divider()
            st.subheader("Results")
            st.dataframe(
                styled_df(disp, val_cols),
                use_container_width=True, hide_index=True,
            )
            st.markdown(
                '<div class="order-note">'
                'Rows = crawled domains, in the order you supplied · '
                'Columns = lines checked, in the order you supplied · '
                'Values = <b>Yes</b> / <b>No</b> / <b>Error</b>'
                '</div>',
                unsafe_allow_html=True,
            )
        except Exception as exc:
            st.warning(f"Pivot build failed ({exc}) - showing flat view.")
            disp = reorder_flat(df_long, domain_order, ordered_keys)
            disp = disp[["Domain", "Line", "Found"]].rename(columns={"Domain": "Crawled Domain"})
            st.dataframe(styled_df(disp, ["Found"]), use_container_width=True, hide_index=True)

        # Fetch-error breakdown
        # groupby ensures one row per domain (no non-unique index), reindex preserves user order
        errs = (df_long[df_long["Found"] == "Error"][["Domain", "_err"]]
                .groupby("Domain", sort=False)["_err"].first()
                .rename("Error reason")
                .reindex(domain_order)
                .dropna()
                .reset_index())
        if not errs.empty:
            with st.expander(f"⚠️ {len(errs)} fetch error(s) - click to expand"):
                st.dataframe(errs, use_container_width=True, hide_index=True)

        st.divider()
        ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
        slug = file_type.replace(".", "_")
        c1, c2 = st.columns(2)
        with c1:
            st.download_button("⬇️ Pivot CSV",
                data=disp.to_csv(index=False).encode(),
                file_name=f"pivot_{slug}_{ts}.csv", mime="text/csv",
                use_container_width=True)
        with c2:
            flat = reorder_flat(df_long, domain_order, ordered_keys)
            flat = flat[["Domain", "Line", "Found"]].rename(columns={"Domain": "Crawled Domain"})
            st.download_button("⬇️ Flat list CSV",
                data=flat.to_csv(index=False).encode(),
                file_name=f"flat_{slug}_{ts}.csv", mime="text/csv",
                use_container_width=True)


# ─────────────────────────────────────────────────────────────────────────────
# TAB 2 - Network Finder
# ─────────────────────────────────────────────────────────────────────────────

with tab2:

    st.markdown(
        "Enter an ad-network domain and relation. "
        "**Leave seller IDs blank** to find every matching entry. "
        "**Upload / paste seller IDs** to check specific ones - "
        "the app checks each domain for lines like "
        "`network, <seller id>, RELATION` built from your IDs."
    )

    nc1, nc2 = st.columns([2, 1])
    with nc1:
        net_input = st.text_input(
            "Ad Network domain", value="xapads.com",
            placeholder="xapads.com", key="net_in",
        )
    with nc2:
        rel_sel = st.selectbox("Relation", ["DIRECT", "RESELLER", "Any"], key="rel_sel")

    st.divider()

    st.markdown("**Seller IDs to check** *(optional)*")
    if saved_exists("seller_ids"):
        st.caption("📌 Loaded saved seller IDs — edit below or change in 💾 Saved Config tab.")
    sid_m = st.radio(
        "sid_m", ["📁 Upload seller IDs", "✏️ Paste seller IDs", "— Find All (no IDs)"],
        index=2, horizontal=True, label_visibility="collapsed", key="sid_m",
    )
    seller_ids_raw = ""
    if sid_m == "📁 Upload seller IDs":
        sid_up = st.file_uploader(
            "seller_ids.txt", type="txt", key="sid_up",
            label_visibility="collapsed",
            help="One seller ID per line (or comma-separated).",
        )
        if sid_up:
            seller_ids_raw = sid_up.read().decode("utf-8", errors="replace")
    elif sid_m == "✏️ Paste seller IDs":
        seller_ids_raw = st.text_area(
            "Seller IDs", value=load_saved("seller_ids"), key="sid_txt", height=130,
            label_visibility="collapsed",
            placeholder="seller_id_001\nseller_id_002\nseller_id_003",
        )

    seller_ids = parse_seller_ids(seller_ids_raw) if seller_ids_raw.strip() else []

    # Mode preview
    if seller_ids:
        rel_label = rel_sel if rel_sel != "Any" else "<any relation>"
        example   = f"{normalize_domain(net_input)}, {seller_ids[0]}, {rel_label}"
        st.markdown(
            f'<span class="mode-tag">🎯 Check Specific IDs - {len(seller_ids)} seller ID(s)</span><br>'
            f'<small style="color:#666">Example line checked: <code>{example}</code> '
            f'(repeated for each ID)</small><br>'
            f'<span class="order-note">Output: Domain | Found (Yes if ANY ID matches) | Details (which ID matched)</span>',
            unsafe_allow_html=True,
        )
    else:
        st.markdown(
            '<span class="mode-tag">🔍 Find All - list every matching seller ID per domain</span><br>'
            '<span class="order-note">Output: Domain | Found | Details (all matching seller IDs)</span>',
            unsafe_allow_html=True,
        )

    if st.button(
        f"🔎 Run Network Finder  [{file_type}]",
        type="primary", key="run_finder", use_container_width=True,
    ):
        if not domain_list:
            st.error("Add domains in the sidebar first.")
        elif not net_input.strip():
            st.error("Enter an ad-network domain.")
        else:
            _net     = normalize_domain(net_input)
            _rel     = "" if rel_sel == "Any" else rel_sel
            _to, _ft = timeout, file_type
            _domains = domain_list[:]

            if seller_ids:
                _sids = seller_ids[:]
                rows = run_parallel(
                    _domains,
                    lambda d: crawl_network_specific(d, _net, _rel, _sids, _to, _ft),
                    workers,
                )
                mode = "specific"
            else:
                rows = run_parallel(
                    _domains,
                    lambda d: crawl_network_all(d, _net, _rel, _to, _ft),
                    workers,
                )
                mode = "all"

            rows = reorder_rows_by_domain(rows, _domains)

            st.session_state.finder_df      = pd.DataFrame(rows)
            st.session_state.finder_meta    = (_net, rel_sel, _ft)
            st.session_state.finder_mode    = mode
            st.session_state.finder_domains = _domains
            st.session_state.finder_nsids   = len(seller_ids)

    # ── Results ───────────────────────────────────────────────────────────────
    if st.session_state.finder_df is not None:
        net_lbl, rel_lbl, ft_lbl = st.session_state.finder_meta or ("", "", "")
        mode         = st.session_state.finder_mode
        domain_order = st.session_state.finder_domains or domain_list

        stale_warn(ft_lbl, file_type)

        df_f    = st.session_state.finder_df
        yes_f   = (df_f["Found"] == "Yes").sum()
        no_f    = (df_f["Found"] == "No").sum()
        err_f   = (df_f["Found"] == "Error").sum()
        total_f = len(df_f)

        st.markdown(tiles_html([
            ("DOMAINS",  f"{total_f:,}", ""),
            ("YES",      f"{yes_f:,}",   "green"),
            ("NO",       f"{no_f:,}",    "red"),
            ("ERRORS",   f"{err_f:,}",   "amber"),
        ]), unsafe_allow_html=True)

        st.divider()
        mode_label = "Specific Seller IDs" if mode == "specific" else "Find All"
        st.subheader(f"Results - {net_lbl} · {rel_lbl}  ({mode_label})")
        if mode == "specific":
            n_ids = st.session_state.finder_nsids or 0
            st.caption(
                f"Checked {n_ids} seller ID(s) per domain. "
                f"**Yes** = at least one matched (see Details for which)."
            )

        fc1, fc2 = st.columns(2)
        with fc1:
            sf = st.multiselect("Filter Found", ["Yes", "No", "Error"],
                                default=["Yes", "No", "Error"], key="ff_unified")
        with fc2:
            dom_f = st.multiselect(
                "Filter domain", df_f["Domain"].tolist(),  # original order, not sorted
                default=[], key="ff_dom_unified",
            )

        view_f = df_f[df_f["Found"].isin(sf)]
        if dom_f:
            view_f = view_f[view_f["Domain"].isin(dom_f)]

        COLS   = ["Domain", "Found", "Details"]
        disp_f = view_f[COLS]

        st.dataframe(
            styled_df(disp_f, ["Found"]),
            use_container_width=True, hide_index=True,
            column_config={
                "Domain":  st.column_config.TextColumn("Domain",  width="medium"),
                "Found":   st.column_config.TextColumn("Found",   width="small"),
                "Details": st.column_config.TextColumn("Details", width="large"),
            },
        )
        st.markdown(
            '<div class="order-note">Rows shown in the order you supplied the domains.</div>',
            unsafe_allow_html=True,
        )
        st.caption(f"Showing {len(view_f):,} of {total_f:,} domains")

        errs_f = (df_f[df_f["Found"] == "Error"][["Domain", "Details"]]
                  .rename(columns={"Details": "Error reason"}))
        if not errs_f.empty:
            with st.expander(f"⚠️ {len(errs_f)} fetch error(s)"):
                st.dataframe(errs_f, use_container_width=True, hide_index=True)

        st.divider()
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        st.download_button(
            "⬇️ Export CSV",
            data=disp_f.to_csv(index=False).encode(),
            file_name=f"finder_{net_lbl}_{mode}_{ts}.csv",
            mime="text/csv", use_container_width=True,
        )


# ─────────────────────────────────────────────────────────────────────────────
# TAB 3 - Combined Export
# ─────────────────────────────────────────────────────────────────────────────

with tab3:

    st.markdown(
        "Enter **both** an ad-network domain **and** the lines you want to check. "
        "One crawl per domain — results land in a single colour-coded Excel: "
        "**Domain → Network Found → Network Details → Line 1 → Line 2 → …**"
    )

    st.divider()

    # ── Section 1: Network ────────────────────────────────────────────────────
    st.markdown("### 1 · Ad Network")
    cn1, cn2 = st.columns([2, 1])
    with cn1:
        comb_net_input = st.text_input(
            "Ad Network domain", value="xapads.com",
            placeholder="xapads.com", key="comb_net",
        )
    with cn2:
        comb_rel_sel = st.selectbox(
            "Relation", ["DIRECT", "RESELLER", "Any"], key="comb_rel",
        )

    st.markdown("**Seller IDs** *(optional — leave blank to list all matching IDs)*")
    if saved_exists("seller_ids"):
        st.caption("📌 Loaded saved seller IDs — edit below or change in 💾 Saved Config tab.")
    comb_sid_m = st.radio(
        "comb_sid_m",
        ["📁 Upload seller IDs", "✏️ Paste seller IDs", "— Find All (no IDs)"],
        index=2, horizontal=True, label_visibility="collapsed", key="comb_sid_m",
    )
    comb_seller_ids_raw = ""
    if comb_sid_m == "📁 Upload seller IDs":
        comb_sid_up = st.file_uploader(
            "seller_ids.txt", type="txt", key="comb_sid_up",
            label_visibility="collapsed",
            help="One seller ID per line (or comma-separated).",
        )
        if comb_sid_up:
            comb_seller_ids_raw = comb_sid_up.read().decode("utf-8", errors="replace")
    elif comb_sid_m == "✏️ Paste seller IDs":
        comb_seller_ids_raw = st.text_area(
            "Seller IDs", value=load_saved("seller_ids"), key="comb_sid_txt", height=100,
            label_visibility="collapsed",
            placeholder="seller_id_001\nseller_id_002\nseller_id_003",
        )
    comb_seller_ids = parse_seller_ids(comb_seller_ids_raw) if comb_seller_ids_raw.strip() else []

    if comb_seller_ids:
        st.caption(f"✔ {len(comb_seller_ids)} seller ID(s) loaded — will check these specifically.")
    else:
        st.caption("No seller IDs supplied — will list every matching seller ID per domain.")

    st.divider()

    # ── Section 2: Lines to check ─────────────────────────────────────────────
    st.markdown("### 2 · Lines to Check")
    comb_lines_save_key = "lines_ads" if file_type == "ads.txt" else "lines_appads"
    if saved_exists(comb_lines_save_key):
        st.caption(
            f"📌 Loaded saved lines for **{file_type}** "
            f"— edit below or change in 💾 Saved Config tab."
        )
    comb_lm = st.radio(
        "comb_lm", ["📁 Upload lines.txt", "✏️ Paste lines"],
        horizontal=True, label_visibility="collapsed", key="comb_lm",
    )
    comb_lines_raw = ""
    if comb_lm == "📁 Upload lines.txt":
        comb_lup = st.file_uploader(
            "lines.txt", type="txt", key="comb_lup",
            label_visibility="collapsed",
            help="One entry per line: domain, seller_id, relation[, tag]",
        )
        if comb_lup:
            comb_lines_raw = comb_lup.read().decode("utf-8", errors="replace")
        else:
            comb_lines_raw = load_saved(comb_lines_save_key)
    else:
        comb_lines_raw = st.text_area(
            "Lines to check", value=load_saved(comb_lines_save_key), key="comb_ltxt", height=150,
            label_visibility="collapsed",
            placeholder=(
                "google.com, pub-6968738577620513, RESELLER, f08c47fec0942fa0\n"
                "smartadserver.com, 5427, RESELLER, 060d053dcf45cbf3\n"
                "video.unrulymedia.com, 906189653, RESELLER"
            ),
        )
    comb_query_list = parse_query_list(comb_lines_raw) if comb_lines_raw.strip() else []
    if comb_query_list:
        st.caption(f"✔ {len(comb_query_list)} line(s) loaded")

    st.divider()

    # ── Run button ────────────────────────────────────────────────────────────
    if st.button(
        f"🚀 Run Combined Export  [{file_type}]",
        type="primary", key="run_combined", use_container_width=True,
    ):
        if not domain_list:
            st.error("Add domains in the sidebar first.")
        elif not comb_net_input.strip():
            st.error("Enter an ad-network domain (Section 1).")
        elif not comb_query_list:
            st.error("Provide at least one line to check (Section 2).")
        else:
            _net     = normalize_domain(comb_net_input)
            _rel     = "" if comb_rel_sel == "Any" else comb_rel_sel
            _sids    = comb_seller_ids[:]
            _ql      = comb_query_list[:]
            _to, _ft, _mf = timeout, file_type, match_fields
            _domains = domain_list[:]

            comb_rows_raw = run_parallel(
                _domains,
                lambda d: crawl_combined(d, _net, _rel, _sids, _ql, _mf, _to, _ft),
                workers,
            )
            comb_rows_raw = reorder_rows_by_domain(comb_rows_raw, _domains)

            net_label  = f"{_net} · {comb_rel_sel}"
            comb_lkeys = [_qkey(q) for q in _ql]

            st.session_state.comb_rows      = comb_rows_raw
            st.session_state.comb_meta      = (net_label, comb_rel_sel, _ft)
            st.session_state.comb_line_keys = comb_lkeys
            st.session_state.comb_domains   = _domains
            st.session_state.comb_nsids     = len(comb_seller_ids)

    # ── Results ───────────────────────────────────────────────────────────────
    if st.session_state.comb_rows is not None:
        c_net_lbl, c_rel_lbl, c_ft_lbl = st.session_state.comb_meta or ("", "", "")
        c_domain_order = st.session_state.comb_domains or domain_list
        c_line_keys    = st.session_state.comb_line_keys or []
        c_rows         = st.session_state.comb_rows

        stale_warn(c_ft_lbl, file_type)

        df_c = pd.DataFrame(c_rows)

        # Metric tiles
        yes_c   = (df_c["Network Found"] == "Yes").sum()
        no_c    = (df_c["Network Found"] == "No").sum()
        err_c   = (df_c["Network Found"] == "Error").sum()
        total_c = len(df_c)

        # Line-level stats across all line columns
        line_cols_present = [k for k in c_line_keys if k in df_c.columns]
        if line_cols_present:
            line_yes = (df_c[line_cols_present] == "Yes").sum().sum()
            line_tot = df_c[line_cols_present].size
            line_pct = f"{round(line_yes / line_tot * 100, 1)}%" if line_tot else "0%"
        else:
            line_yes, line_pct = 0, "0%"

        st.markdown(tiles_html([
            ("DOMAINS",      f"{total_c:,}",  ""),
            ("NETWORK YES",  f"{yes_c:,}",    "green"),
            ("NETWORK NO",   f"{no_c:,}",     "red"),
            ("ERRORS",       f"{err_c:,}",    "amber"),
            ("LINE FOUND %", line_pct,         ""),
        ]), unsafe_allow_html=True)

        st.divider()
        st.subheader(f"Combined Results — {c_net_lbl}")

        n_ids = st.session_state.comb_nsids or 0
        if n_ids:
            st.caption(
                f"Network: checked {n_ids} seller ID(s) per domain. "
                f"**Yes** = at least one matched."
            )

        # Build and display combined dataframe
        display_cols = ["Domain", "Network Found", "Network Details"] + c_line_keys
        df_disp_c = df_c[[col for col in display_cols if col in df_c.columns]]

        color_target_cols = ["Network Found"] + [k for k in c_line_keys if k in df_c.columns]
        st.dataframe(
            styled_df(df_disp_c, color_target_cols),
            use_container_width=True, hide_index=True,
            column_config={
                "Domain":          st.column_config.TextColumn("Domain",          width="medium"),
                "Network Found":   st.column_config.TextColumn("Network Found",   width="small"),
                "Network Details": st.column_config.TextColumn("Network Details", width="large"),
            },
        )
        st.markdown(
            '<div class="order-note">'
            'Col 1: Crawled domains in your supplied order · '
            'Col 2–3: Network check · '
            'Remaining cols: one per line checked · '
            'Values: <b>Yes</b> / <b>No</b> / <b>Error</b>'
            '</div>',
            unsafe_allow_html=True,
        )
        st.caption(f"Showing {total_c:,} domain(s) · {len(c_line_keys)} line(s) checked")

        # Error breakdown
        errs_c = df_c[df_c["Network Found"] == "Error"][["Domain", "Network Details"]].rename(
            columns={"Network Details": "Error reason"}
        )
        if not errs_c.empty:
            with st.expander(f"⚠️ {len(errs_c)} fetch error(s)"):
                st.dataframe(errs_c, use_container_width=True, hide_index=True)

        st.divider()
        ts_c = datetime.now().strftime("%Y%m%d_%H%M%S")
        net_slug = c_net_lbl.replace(" ", "_").replace("·", "").replace("/", "-")

        excel_bytes = build_combined_excel(c_rows, c_net_lbl, c_line_keys)
        st.download_button(
            "⬇️ Download Combined Excel (.xlsx)",
            data=excel_bytes,
            file_name=f"combined_{net_slug}_{ts_c}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

# ─────────────────────────────────────────────────────────────────────────────
# TAB 4 - Saved Config
# ─────────────────────────────────────────────────────────────────────────────

with tab4:

    st.markdown(
        "Save the things you crawl with **every day** here once - they'll be "
        "pre-loaded automatically into the sidebar and tabs next time you open "
        "the app, so you don't need to re-paste your sheet each time.\n\n"
        "You can still **edit or replace** anything in the other tabs at any "
        "time - that won't affect what's saved here unless you save again."
    )

    if PERSISTENCE_WARNING:
        st.warning(PERSISTENCE_WARNING, icon="⚠️")

    st.divider()

    def _saved_config_section(key: str, height: int = 150, placeholder: str = ""):
        label = SAVED_LABELS[key]
        st.markdown(f"### {label}")

        current = load_saved(key)
        if current.strip():
            n_lines = len([l for l in current.splitlines() if l.strip()])
            st.caption(f"✅ Currently saved: {n_lines} non-empty line(s).")
        else:
            st.caption("— Nothing saved yet —")

        mode = st.radio(
            f"{key}_mode", ["✏️ Paste", "📁 Upload"],
            horizontal=True, label_visibility="collapsed", key=f"{key}_save_mode",
            disabled=not PERSISTENCE_OK,
        )

        new_content = ""
        if mode == "📁 Upload":
            up = st.file_uploader(
                f"{key}.txt", type="txt", key=f"{key}_save_up",
                label_visibility="collapsed", disabled=not PERSISTENCE_OK,
            )
            if up:
                new_content = up.read().decode("utf-8", errors="replace")
        else:
            new_content = st.text_area(
                label, value=current, height=height, key=f"{key}_save_txt",
                label_visibility="collapsed", placeholder=placeholder,
                disabled=not PERSISTENCE_OK,
            )

        bc1, bc2, bc3 = st.columns([1, 1, 3])
        with bc1:
            if st.button("💾 Save", key=f"{key}_save_btn",
                          use_container_width=True, disabled=not PERSISTENCE_OK):
                if save_saved(key, new_content):
                    st.success("Saved.")
                    safe_rerun()
                else:
                    st.error("Couldn't save — persistent storage isn't available right now.")
        with bc2:
            if st.button("🗑️ Clear", key=f"{key}_clear_btn",
                          use_container_width=True, disabled=not PERSISTENCE_OK):
                if clear_saved(key):
                    st.success("Cleared.")
                    safe_rerun()
                else:
                    st.error("Couldn't clear — persistent storage isn't available right now.")

        if current.strip():
            with st.expander("Preview currently saved content", expanded=False):
                st.code(current, language="text")

        st.divider()

    _saved_config_section(
        "domains", height=180,
        placeholder="apps.mxplayer.in\nbattleprime.com\n…",
    )
    _saved_config_section(
        "lines_ads", height=180,
        placeholder=(
            "google.com, pub-6968738577620513, RESELLER, f08c47fec0942fa0\n"
            "smartadserver.com, 5427, RESELLER, 060d053dcf45cbf3"
        ),
    )
    _saved_config_section(
        "lines_appads", height=180,
        placeholder=(
            "google.com, pub-6968738577620513, RESELLER, f08c47fec0942fa0\n"
            "smartadserver.com, 5427, RESELLER, 060d053dcf45cbf3"
        ),
    )
    _saved_config_section(
        "seller_ids", height=130,
        placeholder="seller_id_001\nseller_id_002\nseller_id_003",
    )
